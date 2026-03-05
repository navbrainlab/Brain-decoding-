#!/usr/bin/env python3
import argparse
import json
import os
import subprocess
import sys
import numpy as np
from typing import List
from typing import Optional


def find_latest_dir(root: str) -> str:
    if not os.path.isdir(root):
        raise FileNotFoundError(f"Output path does not exist: {root}")
    subdirs = [os.path.join(root, d) for d in os.listdir(root) if os.path.isdir(os.path.join(root, d))]
    if not subdirs:
        raise FileNotFoundError(f"No experiment directories found under: {root}")
    # Sort by name (timestamps like YYYYMMDD_HHMMSS sort correctly lexicographically)
    subdirs.sort()
    return subdirs[-1]


def run_cmd(cmd: List[str]) -> None:
    print(f"[RUN] {' '.join(cmd)}")
    result = subprocess.run(cmd)
    if result.returncode != 0:
        raise RuntimeError(f"Command failed with exit code {result.returncode}: {' '.join(cmd)}")


def _load_json(path: str) -> Optional[dict]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _latest_adapt_args(checkpoint_root: str) -> Optional[tuple]:
    # Find latest args_YYYYMMDD_HHMMSS.json created by adapt.py
    try:
        candidates = [
            f for f in os.listdir(checkpoint_root)
            if f.startswith("args_") and f.endswith(".json")
        ]
        if not candidates:
            return None
        candidates.sort()
        latest_file = candidates[-1]
        args_obj = _load_json(os.path.join(checkpoint_root, latest_file))
        # Extract timestamp between prefix and extension
        adapt_id = latest_file[len("args_"):-len(".json")]
        return args_obj, adapt_id
    except Exception:
        return None


def _get_final_accuracy(checkpoint_root: str, target_date: str) -> Optional[float]:
    npz_path = os.path.join(checkpoint_root, f"adaptation_results_{target_date}.npz")
    if not os.path.isfile(npz_path):
        return None
    try:
        with np.load(npz_path, allow_pickle=True) as npz:
            acc = npz.get("accuracy", None)
            if acc is None:
                return None
            acc = acc.flatten()
            return float(acc[-1]) if acc.size > 0 else None
    except Exception:
        return None


def _mmdd_from_date(date_str: str) -> Optional[str]:
    # Try to extract MMDD from typical formats: '20250326' -> '0326', '0326' -> '0326'
    digits = ''.join(ch for ch in date_str if ch.isdigit())
    if len(digits) >= 8:
        return digits[4:8]
    if len(digits) == 4:
        return digits
    return None


def _npz_scalar(path: str, key: str) -> Optional[float]:
    if not os.path.isfile(path):
        return None
    try:
        with np.load(path, allow_pickle=True) as npz:
            val = npz.get(key, None)
            if val is None:
                return None
            arr = np.array(val).flatten()
            # If it's a vector (e.g., accuracy curve), take last
            return float(arr[-1]) if arr.size > 0 else None
    except Exception:
        return None


def _mf_source_model_meta(mf_root: str, source_date: str) -> dict:
    meta = {}
    model_npz = os.path.join(mf_root, f"source_model_{source_date}.npz")
    if os.path.isfile(model_npz):
        try:
            with np.load(model_npz, allow_pickle=True) as npz:
                # Shapes to infer n_class
                M_spd = npz.get("M_spd", None)
                M_deep = npz.get("M_deep", None)
                if M_spd is not None:
                    meta["n_class"] = int(M_spd.shape[1])
                elif M_deep is not None:
                    meta["n_class"] = int(M_deep.shape[1])
                meta["alignment"] = str(npz.get("alignment", ""))
                meta["similarity"] = str(npz.get("similarity", ""))
                meta["k"] = int(npz.get("k", 0))
                meta["lambda_hg"] = float(npz.get("lambda_hg", 0.0))
                meta["miu"] = float(npz.get("miu", 0.0))
                meta["eta"] = float(npz.get("eta", 0.0))
        except Exception:
            pass
    return meta


def _mf_latest_adapt_log_id(mf_root: str) -> Optional[str]:
    try:
        candidates = [f for f in os.listdir(mf_root) if f.startswith("adapt_") and f.endswith(".log")]
        if not candidates:
            return None
        candidates.sort()
        latest = candidates[-1]
        # adapt_{exp_id}.log
        return latest[len("adapt_"):-len(".log")]
    except Exception:
        return None


def main():
    parser = argparse.ArgumentParser(description="Run training then adaptation sequentially.")
    parser.add_argument("--train-script", default="train.py", help="Path to train.py")
    parser.add_argument("--adapt-script", default="adapt.py", help="Path to adapt.py")
    parser.add_argument(
        "--output-path",
        default="/media/ubuntu/Storage1/ecog_data/OTTA_results",
        help="Root directory where train.py writes experiment timestamp folders",
    )
    parser.add_argument(
        "--target-dates",
        default='20250710,20250711,20250804,20250814,20250825',
        help="Comma-separated target date(s) to use for adaptation (e.g., 'E' or '20250329,20250331')",
    )
    # Common overrides for both train.py and adapt.py
    parser.add_argument("--seed", type=int, default=37, help="Random seed to pass to both train.py and adapt.py")
    parser.add_argument("--source-dates", default='20250701_20250702_20250703_20250704_20250707_20250708_20250709', help="Comma-separated source date(s) to pass to both scripts (e.g., 'T' or '20250325_20250326_20250327')")
    parser.add_argument("--data-path", default='/media/ubuntu/Storage1/ecog_data/new_day_data/C07_BDY-S01_CF-8Fr/daily/beida_s01_angle_filtered.h5', help="Dataset path to pass to both train.py and adapt.py (e.g., 'bcic_1' or '/path/to/daily_bdy')")
    # Global consistency controls
    parser.add_argument("--alignment", default="Riemannian", help="Global alignment value to use across all scripts that accept it (e.g., 'Riemannian' or 'Euclidean')")
    parser.add_argument("--similarity", default="tangent_DM", help="Global similarity to use for multi_feature_hypergraph_*.py (e.g., 'SPD', 'tangent_cosine', 'tangent_DM', 'SPD_DM', 'cosine', 'DM')")
    # Optional: allow passing extra args to train.py without parsing them here
    parser.add_argument(
        "--train-extra",
        default="",
        help="Extra arguments forwarded to train.py (single string)",
    )
    # Optional: allow passing extra args to adapt.py without parsing them here
    parser.add_argument(
        "--adapt-extra",
        default="",
        help="Extra arguments forwarded to adapt.py (single string)",
    )
    # Optional: allow passing extra args to multi_feature_hypergraph_train.py
    parser.add_argument(
        "--multi-feature-extra",
        default="",
        help="Extra arguments forwarded to multi_feature_hypergraph_train.py (single string)",
    )
    # Output root for multi-feature hypergraph training results
    parser.add_argument(
        "--multi-feature-output-path",
        default="/media/ubuntu/Storage1/ecog_data/Multi-feature_results",
        help="Root directory where multi_feature_hypergraph_train.py writes timestamp folders",
    )
    # Output root for Riemannian MDM training results
    parser.add_argument(
        "--mdm-output-path",
        default="/media/ubuntu/Storage1/ecog_data/MDM_results",
        help="Root directory where Riemannian_MDM_train.py writes timestamp folders",
    )
    # Output root for resemble adaptation results
    parser.add_argument(
        "--resemble-output-path",
        default="/media/ubuntu/Storage1/ecog_data/resemble_results",
        help="Root directory where resemble_adapt.py writes timestamp folders",
    )
    # Optional: extra args for multi_feature_hypergraph_adapt.py
    parser.add_argument(
        "--multi-feature-adapt-extra",
        default="",
        help="Extra arguments forwarded to multi_feature_hypergraph_adapt.py (single string)",
    )
    args = parser.parse_args()

    # 1) Run training
    train_cmd = [sys.executable, args.train_script]
    # Forward seed and source dates
    if args.seed is not None:
        train_cmd.extend(["--seed", str(args.seed)])
    if args.source_dates:
        src_list = [s for s in args.source_dates.split(",") if s]
        # nargs='+' expects values after the flag
        train_cmd.append("--source_dates")
        train_cmd.extend(src_list)
    if args.data_path:
        train_cmd.extend(["--data_path", args.data_path])
    if args.train_extra:
        # Split by spaces; caller is responsible for proper quoting
        train_cmd.extend(args.train_extra.split())
    # Enforce global alignment if provided
    if args.alignment:
        train_cmd.extend(["--alignment", args.alignment])
    run_cmd(train_cmd)

    # 2) Find latest experiment directory under output path
    latest_dir = find_latest_dir(args.output_path)
    print(f"[INFO] Latest experiment directory: {latest_dir}")

    # 3) Locate and update adapt_config.json with target_dates
    cfg_path = os.path.join(latest_dir, "adapt_config.json")
    if not os.path.isfile(cfg_path):
        raise FileNotFoundError(f"adapt_config.json not found at {cfg_path}")

    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    target_dates_list = [s for s in args.target_dates.split(",") if s]
    cfg["target_dates"] = target_dates_list

    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)
    print(f"[INFO] Updated target_dates in {cfg_path}: {target_dates_list}")

    # # 4) Run adaptation with the config
    # adapt_cmd = [sys.executable, args.adapt_script, "--config", cfg_path]
    # # Forward seed and source dates to override config if provided
    # if args.seed is not None:
    #     adapt_cmd.extend(["--seed", str(args.seed)])
    # if args.source_dates:
    #     src_list = [s for s in args.source_dates.split(",") if s]
    #     adapt_cmd.append("--source_dates")
    #     adapt_cmd.extend(src_list)
    # # Forward target dates via CLI as well (in addition to updating config)
    # if args.target_dates:
    #     tgt_list = [s for s in args.target_dates.split(",") if s]
    #     adapt_cmd.append("--target_dates")
    #     adapt_cmd.extend(tgt_list)
    # if args.data_path:
    #     adapt_cmd.extend(["--data_path", args.data_path])
    # if args.alignment:
    #     adapt_cmd.extend(["--alignment", args.alignment])
    # if args.adapt_extra:
    #     # Split by spaces; caller is responsible for proper quoting
    #     adapt_cmd.extend(args.adapt_extra.split())
    # run_cmd(adapt_cmd)

    # 5) Run multi_feature_hypergraph_train.py using artifacts from training
    # Read training args to get authoritative source_dates used by train.py
    train_args_path = os.path.join(latest_dir, "args.json")
    if not os.path.isfile(train_args_path):
        raise FileNotFoundError(f"Training args.json not found at {train_args_path}")
    with open(train_args_path, "r", encoding="utf-8") as f:
        train_args_obj = json.load(f)

    mf_cmd = [sys.executable, "multi_feature_hypergraph_train.py"]
    # checkpoint_root should point to latest training directory containing best_model_* and args.json
    mf_cmd.extend(["--checkpoint_root", latest_dir])
    # ensure multi-feature outputs are written under the specified root
    mf_cmd.extend(["--output_path", args.multi_feature_output_path])
    # Forward seed and data_path if provided
    if args.seed is not None:
        mf_cmd.extend(["--seed", str(args.seed)])
    if args.data_path:
        mf_cmd.extend(["--data_path", args.data_path])
    # Forward source_dates from training for consistency (nargs-style)
    # Prefer explicit args.source_dates; otherwise fall back to training args.json
    if args.source_dates:
        real_source_dates = [s for s in args.source_dates.split(",") if s]
    else:
        real_source_dates = train_args_obj.get("source_dates", []) or []
        if isinstance(real_source_dates, str):
            real_source_dates = [s for s in real_source_dates.split(",") if s]
        elif not isinstance(real_source_dates, list):
            try:
                real_source_dates = list(real_source_dates)
            except Exception:
                real_source_dates = []
    if real_source_dates:
        mf_cmd.append("--source_dates")
        mf_cmd.extend(real_source_dates)
    # Indices path: use the training history file under latest_dir tied to the first source date
    first_src = real_source_dates[0] if real_source_dates else None
    if first_src:
        indices_path = os.path.join(latest_dir, f"training_history_{first_src}.json")
        if os.path.isfile(indices_path):
            mf_cmd.extend(["--indices_path", indices_path])
    # Allow user-specified extras
    if args.multi_feature_extra:
        mf_cmd.extend(args.multi_feature_extra.split())
    # Enforce global alignment/similarity if provided (append last to take precedence)
    if args.alignment:
        mf_cmd.extend(["--alignment", args.alignment])
    if args.similarity:
        mf_cmd.extend(["--similarity", args.similarity])

    run_cmd(mf_cmd)

    # 6) Run Riemannian_MDM_train.py using artifacts from training
    mdm_cmd = [sys.executable, "Riemannian_MDM_train.py"]
    # Ensure outputs are written under the specified root
    mdm_cmd.extend(["--output_path", args.mdm_output_path])
    # Prefer explicit runner inputs; fallback to training args
    # Source dates
    if args.source_dates:
        mdm_source_dates = [s for s in args.source_dates.split(",") if s]
    else:
        sd = train_args_obj.get("source_dates", []) or []
        if isinstance(sd, str):
            mdm_source_dates = [s for s in sd.split(",") if s]
        elif isinstance(sd, list):
            mdm_source_dates = sd
        else:
            try:
                mdm_source_dates = list(sd)
            except Exception:
                mdm_source_dates = []
    if mdm_source_dates:
        mdm_cmd.append("--source_dates")
        mdm_cmd.extend(mdm_source_dates)
    # Data path
    if args.data_path:
        mdm_data_path = args.data_path
    else:
        mdm_data_path = train_args_obj.get("data_path", "")
    if mdm_data_path:
        mdm_cmd.extend(["--data_path", mdm_data_path])
    # Seed
    if args.seed is not None:
        mdm_cmd.extend(["--seed", str(args.seed)])
    # Alignment (prefer global runner arg, else from train args)
    mdm_alignment = args.alignment if args.alignment else train_args_obj.get("alignment")
    if mdm_alignment:
        mdm_cmd.extend(["--alignment", mdm_alignment])
    # Indices path: derive from training artifacts
    mdm_first_src = mdm_source_dates[0] if mdm_source_dates else None
    if mdm_first_src:
        mdm_indices_path = os.path.join(latest_dir, f"training_history_{mdm_first_src}.json")
        if os.path.isfile(mdm_indices_path):
            mdm_cmd.extend(["--indices_path", mdm_indices_path])
    run_cmd(mdm_cmd)

    # # 7) Run Riemannian_MDM_adapt.py using latest MDM outputs and training artifacts
    mdm_latest_dir = find_latest_dir(args.mdm_output_path)
    print(f"[INFO] Latest MDM directory: {mdm_latest_dir}")

    # Compose config to ensure list-typed args parsed correctly
    mdm_target_dates = [s for s in args.target_dates.split(",") if s]

    # mdm_adapt_cfg = {
    #     "class_means_root": mdm_latest_dir,
    #     "indices_root": latest_dir,
    #     "source_dates": mdm_source_dates if 'mdm_source_dates' in locals() else [],
    #     "target_dates": mdm_target_dates,
    #     "data_path": mdm_data_path if 'mdm_data_path' in locals() else (args.data_path or train_args_obj.get("data_path", "")),
    #     # alignment/update/buffer params can stay defaults unless overridden via extras
    # }

    # mdm_adapt_cfg_path = os.path.join(mdm_latest_dir, "mdm_adapt_config.json")
    # with open(mdm_adapt_cfg_path, "w", encoding="utf-8") as f:
    #     json.dump(mdm_adapt_cfg, f, indent=2)
    # print(f"[INFO] Wrote MDM adapt config: {mdm_adapt_cfg_path}")

    # mdm_adapt_cmd = [sys.executable, "Riemannian_MDM_adapt.py", "--config", mdm_adapt_cfg_path]
    # # Forward simple overrides
    # if args.seed is not None:
    #     mdm_adapt_cmd.extend(["--seed", str(args.seed)])
    # if args.data_path:
    #     mdm_adapt_cmd.extend(["--data_path", args.data_path])
    # if args.alignment:
    #     mdm_adapt_cmd.extend(["--alignment", args.alignment])
    # run_cmd(mdm_adapt_cmd)

    # 8) Run resemble_adapt.py using latest train and MDM outputs
    resemble_cfg = {
        "output_root": args.resemble_output_path,
        "checkpoint_root": latest_dir,
        "class_means_root": mdm_latest_dir,
        "indices_root": latest_dir,
        "source_dates": mdm_source_dates if 'mdm_source_dates' in locals() else (train_args_obj.get("source_dates", []) or []),
        "target_dates": mdm_target_dates,
        "data_path": mdm_data_path if 'mdm_data_path' in locals() else (args.data_path or train_args_obj.get("data_path", "")),
    }
    resemble_cfg_path = os.path.join(latest_dir, "resemble_adapt_config.json")
    with open(resemble_cfg_path, "w", encoding="utf-8") as f:
        json.dump(resemble_cfg, f, indent=2)
    print(f"[INFO] Wrote resemble adapt config: {resemble_cfg_path}")

    resemble_cmd = [sys.executable, "resemble_adapt.py", "--config", resemble_cfg_path]
    if args.seed is not None:
        resemble_cmd.extend(["--seed", str(args.seed)])
    if args.data_path:
        resemble_cmd.extend(["--data_path", args.data_path])
    if args.alignment:
        resemble_cmd.extend(["--alignment", args.alignment])
    run_cmd(resemble_cmd)

    # 9) Run multi_feature_hypergraph_adapt.py using training and multi-feature outputs
    # Discover latest multi-feature output directory
    mf_latest_dir = find_latest_dir(args.multi_feature_output_path)
    print(f"[INFO] Latest multi-feature hypergraph directory: {mf_latest_dir}")
    # Discover latest resemble output directory
    res_latest_dir = find_latest_dir(args.resemble_output_path)
    print(f"[INFO] Latest resemble directory: {res_latest_dir}")
    
    # Build a config JSON to ensure list-typed args (source/target dates) are parsed correctly
    # Prefer explicit runner inputs for source/target dates; fallback to training args
    if args.source_dates:
        mf_source_dates = [s for s in args.source_dates.split(",") if s]
    else:
        sd = train_args_obj.get("source_dates", []) or []
        if isinstance(sd, str):
            mf_source_dates = [s for s in sd.split(",") if s]
        elif isinstance(sd, list):
            mf_source_dates = sd
        else:
            try:
                mf_source_dates = list(sd)
            except Exception:
                mf_source_dates = []
    mf_target_dates = [s for s in args.target_dates.split(",") if s]

    # Determine indices_root so indices match those saved by adapt.py's record_dict
    # Prefer the per-target adaptation results (adaptation_results_{target}.npz) saved by adapt.py under checkpoint_root.
    first_src = mf_source_dates[0] if mf_source_dates else None

    mf_adapt_cfg = {
        "checkpoint_root": latest_dir,
        "source_hypergraph_root": mf_latest_dir,
        "indices_root": res_latest_dir,
        "source_dates": mf_source_dates,
        "target_dates": mf_target_dates,
        "data_path": args.data_path if args.data_path else train_args_obj.get("data_path", ""),
    }
    mf_adapt_cfg_path = os.path.join(mf_latest_dir, "mf_adapt_config.json")
    with open(mf_adapt_cfg_path, "w", encoding="utf-8") as f:
        json.dump(mf_adapt_cfg, f, indent=2)
    print(f"[INFO] Wrote multi-feature adapt config: {mf_adapt_cfg_path}")

    mf_adapt_cmd = [sys.executable, "multi_feature_hypergraph_adapt.py", "--config", mf_adapt_cfg_path]
    # Forward simple overrides
    if args.seed is not None:
        mf_adapt_cmd.extend(["--seed", str(args.seed)])
    if args.data_path:
        mf_adapt_cmd.extend(["--data_path", args.data_path])
    if args.multi_feature_adapt_extra:
        mf_adapt_cmd.extend(args.multi_feature_adapt_extra.split())
    run_cmd(mf_adapt_cmd)

    # 10) Write multi-feature XLSX report (train + adapt)
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        Workbook = None
        load_workbook = None

    mf_headers = [
        "Train ID",
        "Data Version",
        "Source date",
        "Source domain number",
        "n_class",
        "seed",
        "alignment",
        "similarity",
        "k",
        "lambda",
        "miu",
        "eta",
        "OTTA Checkpoint",
        "Multi-layer feature",
        "Validation accuracy \n(Deep)",
        "Validation accuracy\n (SPD)",
        "Validation accuracy",
        "Adapt ID",
        "Target date",
        "buffer_size",
        "buffer_weighting",
        "buffer_decay",
        "0326 accuracy\n (deep)",
        "0326 accuracy\n (spd)",
        "0326 accuracy",
        "0327 accuracy\n (deep)",
        "0327 accuracy\n (spd)",
        "0327 accuracy",
        "0329 accuracy\n (deep)",
        "0329 accuracy\n (spd)",
        "0329 accuracy",
        "0331 accuracy\n (deep)",
        "0331 accuracy\n (spd)",
        "0331 accuracy",
        "0401 accuracy\n (deep)",
        "0401 accuracy\n (spd)",
        "0401 accuracy",
        "E accuracy",
        "0710 accuracy\n (deep)",
        "0710 accuracy\n (spd)",
        "0710 accuracy",
        "0711 accuracy\n (deep)",
        "0711 accuracy\n (spd)",
        "0711 accuracy",
        "0804 accuracy\n (deep)",
        "0804 accuracy\n (spd)",
        "0804 accuracy",
        "0814 accuracy\n (deep)",
        "0814 accuracy\n (spd)",
        "0814 accuracy",
        "0825 accuracy\n (deep)",
        "0825 accuracy\n (spd)",
        "0825 accuracy",        
        "Avg test accuracy\n(Deep)",
        "Avg test accuracy\n(SPD)",
        "Avg test accuracy",
    ]

    # Metadata
    mf_train_id = os.path.basename(mf_latest_dir)
    mf_data_version = args.data_path or train_args_obj.get("data_path", "")
    mf_source_dates = mf_source_dates if 'mf_source_dates' in locals() else (train_args_obj.get("source_dates", []) or [])
    if isinstance(mf_source_dates, str):
        mf_source_dates = [s for s in mf_source_dates.split(',') if s]
    mf_source_num = len(mf_source_dates)
    mf_first_src = mf_source_dates[0] if mf_source_dates else None
    mf_seed = args.seed if args.seed is not None else train_args_obj.get("seed", "")
    mf_meta = _mf_source_model_meta(mf_latest_dir, mf_first_src or '')
    mf_alignment = mf_meta.get("alignment", "")
    mf_similarity = mf_meta.get("similarity", "")
    mf_k = mf_meta.get("k", "")
    mf_lambda = mf_meta.get("lambda_hg", "")
    mf_miu = mf_meta.get("miu", "")
    mf_eta = mf_meta.get("eta", "")
    mf_checkpoint = latest_dir
    mf_multi_layer = False

    # Validation accuracies from MF train npz
    val_npz = os.path.join(mf_latest_dir, f"source_validation_results_{mf_first_src}.npz") if mf_first_src else ''
    val_acc_deep = _npz_scalar(val_npz, "deep_accuracy") if val_npz else None
    val_acc_spd = _npz_scalar(val_npz, "spd_accuracy") if val_npz else None
    val_acc_total = _npz_scalar(val_npz, "source_accuracy") if val_npz else None

    # MF adapt info
    mf_adapt_id = _mf_latest_adapt_log_id(mf_latest_dir) or ''
    mf_tgt_dates = mf_target_dates
    # Buffer params: defaults (unless user overrides via CLI extras, which we don't parse here)
    mf_buffer_size = 32
    mf_buffer_weighting = 'uniform'
    mf_buffer_decay = 2.0

    # Per-target final accuracies (deep/spd/total) from MF adapt results
    mmdd_cols_deep = {
        "0326": "0326 accuracy\n (deep)",
        "0327": "0327 accuracy\n (deep)",
        "0329": "0329 accuracy\n (deep)",
        "0331": "0331 accuracy\n (deep)",
        "0401": "0401 accuracy\n (deep)",
        "0710": "0710 accuracy\n (deep)",
        "0711": "0711 accuracy\n (deep)",
        "0804": "0804 accuracy\n (deep)",
        "0814": "0814 accuracy\n (deep)",
        "0825": "0825 accuracy\n (deep)",
    }
    mmdd_cols_spd = {
        "0326": "0326 accuracy\n (spd)",
        "0327": "0327 accuracy\n (spd)",
        "0329": "0329 accuracy\n (spd)",
        "0331": "0331 accuracy\n (spd)",
        "0401": "0401 accuracy\n (spd)",
        "0710": "0710 accuracy\n (spd)",
        "0711": "0711 accuracy\n (spd)",
        "0804": "0804 accuracy\n (spd)",
        "0814": "0814 accuracy\n (spd)",
        "0825": "0825 accuracy\n (spd)",
    }
    mmdd_cols_total = {
        "0326": "0326 accuracy",
        "0327": "0327 accuracy",
        "0329": "0329 accuracy",
        "0331": "0331 accuracy",
        "0401": "0401 accuracy",
        "E": "E accuracy",
        "0710": "0710 accuracy",
        "0711": "0711 accuracy",
        "0804": "0804 accuracy",
        "0814": "0814 accuracy",
        "0825": "0825 accuracy",
    }
    per_deep = {}
    per_spd = {}
    per_total = {}
    deep_vals = []
    spd_vals = []
    total_vals = []
    for td in mf_tgt_dates:
        npz_path = os.path.join(mf_latest_dir, f"adaptation_results_{td}.npz")
        deep_final = _npz_scalar(npz_path, "deep_accuracy")
        spd_final = _npz_scalar(npz_path, "spd_accuracy")
        total_final = _npz_scalar(npz_path, "accuracy")
        if deep_final is not None:
            deep_vals.append(deep_final)
        if spd_final is not None:
            spd_vals.append(spd_final)
        if total_final is not None:
            total_vals.append(total_final)
        mmdd = _mmdd_from_date(td)
        if not mmdd and str(td).upper() == 'E':
            mmdd = 'E'
        if mmdd:
            if mmdd in mmdd_cols_deep and deep_final is not None:
                per_deep[mmdd_cols_deep[mmdd]] = deep_final
            if mmdd in mmdd_cols_spd and spd_final is not None:
                per_spd[mmdd_cols_spd[mmdd]] = spd_final
            if mmdd in mmdd_cols_total and total_final is not None:
                per_total[mmdd_cols_total[mmdd]] = total_final

    avg_deep = sum(deep_vals) / len(deep_vals) if deep_vals else None
    avg_spd = sum(spd_vals) / len(spd_vals) if spd_vals else None
    avg_total = sum(total_vals) / len(total_vals) if total_vals else None

    mf_row = [
        mf_train_id,
        mf_data_version,
        ','.join(mf_source_dates),
        mf_source_num,
        mf_meta.get("n_class", ""),
        mf_seed,
        mf_alignment,
        mf_similarity,
        mf_k,
        mf_lambda,
        mf_miu,
        mf_eta,
        mf_checkpoint,
        mf_multi_layer,
        val_acc_deep,
        val_acc_spd,
        val_acc_total,
        mf_adapt_id,
        ','.join(mf_tgt_dates),
        mf_buffer_size,
        mf_buffer_weighting,
        mf_buffer_decay,
        per_deep.get("0326 accuracy\n (deep)", None),
        per_spd.get("0326 accuracy\n (spd)", None),
        per_total.get("0326 accuracy", None),
        per_deep.get("0327 accuracy\n (deep)", None),
        per_spd.get("0327 accuracy\n (spd)", None),
        per_total.get("0327 accuracy", None),
        per_deep.get("0329 accuracy\n (deep)", None),
        per_spd.get("0329 accuracy\n (spd)", None),
        per_total.get("0329 accuracy", None),
        per_deep.get("0331 accuracy\n (deep)", None),
        per_spd.get("0331 accuracy\n (spd)", None),
        per_total.get("0331 accuracy", None),
        per_deep.get("0401 accuracy\n (deep)", None),
        per_spd.get("0401 accuracy\n (spd)", None),
        per_total.get("0401 accuracy", None),
        per_total.get("E accuracy", None),
        per_deep.get("0710 accuracy\n (deep)", None),
        per_spd.get("0710 accuracy\n (spd)", None),
        per_total.get("0710 accuracy", None),
        per_deep.get("0711 accuracy\n (deep)", None),
        per_spd.get("0711 accuracy\n (spd)", None),
        per_total.get("0711 accuracy", None),
        per_deep.get("0804 accuracy\n (deep)", None),
        per_spd.get("0804 accuracy\n (spd)", None),
        per_total.get("0804 accuracy", None),
        per_deep.get("0814 accuracy\n (deep)", None),
        per_spd.get("0814 accuracy\n (spd)", None),
        per_total.get("0814 accuracy", None),
        per_deep.get("0825 accuracy\n (deep)", None),
        per_spd.get("0825 accuracy\n (spd)", None),
        per_total.get("0825 accuracy", None),
        avg_deep,
        avg_spd,
        avg_total,
    ]

    mf_report_path = os.path.join(args.multi_feature_output_path, "mf_summary.xlsx")
    try:
        if Workbook is None or load_workbook is None:
            print(f"[WARN] openpyxl not available; skipped MF XLSX report at {mf_report_path}")
        else:
            if os.path.isfile(mf_report_path):
                wb = load_workbook(mf_report_path)
                ws = wb.active
                ws.append(mf_row)
                wb.save(mf_report_path)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(mf_headers)
                ws.append(mf_row)
                wb.save(mf_report_path)
            print(f"[INFO] Wrote/updated MF XLSX report: {mf_report_path}")
    except Exception as e:
        print(f"[WARN] Failed to write MF XLSX report: {e}")

    # 11) Write MDM XLSX report (train + adapt)
    # Aggregate Riemannian_MDM_train.py validation accuracy and Riemannian_MDM_adapt.py per-target final accuracies
    try:
        mdm_headers = [
            "Train ID",
            "Data Version",
            "Source date",
            "Source domain number",
            "n_class",
            "seed",
            "Alignment",
            "Validation accuracy",
            "Adapt ID",
            "Target date",
            "Alignment",
            "buffer_size",
            "buffer_weighting",
            "buffer_decay",
            "0326 accuracy",
            "0327 accuracy",
            "0329 accuracy",
            "0331 accuracy",
            "0401 accuracy",
            "E accuracy",
            "0710 accuracy",
            "0711 accuracy",
            "0804 accuracy",
            "0814 accuracy",
            "0825 accuracy",
            "Avg test accuracy",
        ]

        mdm_train_id = os.path.basename(mdm_latest_dir)
        mdm_args_path = os.path.join(mdm_latest_dir, 'args.json')
        mdm_args_obj = _load_json(mdm_args_path) or {}

        mdm_data_version = args.data_path or mdm_args_obj.get("data_path", "")
        if args.source_dates:
            mdm_source_dates_local = [s for s in args.source_dates.split(',') if s]
        else:
            mdm_source_dates_local = mdm_args_obj.get("source_dates", []) or []
        if isinstance(mdm_source_dates_local, str):
            mdm_source_dates_local = [s for s in mdm_source_dates_local.split(',') if s]
        mdm_source_num = len(mdm_source_dates_local)
        mdm_first_src_local = mdm_source_dates_local[0] if mdm_source_dates_local else ''

        # n_class from class_means npz
        n_class = ''
        cm_npz_path = os.path.join(mdm_latest_dir, f"class_means_source_{mdm_first_src_local}.npz") if mdm_first_src_local else ''
        if cm_npz_path and os.path.isfile(cm_npz_path):
            try:
                with np.load(cm_npz_path) as npz:
                    cm = npz.get('class_means', None)
                    if cm is not None:
                        n_class = cm.shape[0]
            except Exception:
                n_class = ''

        mdm_seed_local = args.seed if args.seed is not None else mdm_args_obj.get("seed", "")
        mdm_alignment_train = mdm_args_obj.get("alignment", "")

        # Parse validation accuracy from latest train log
        mdm_val_acc = None
        try:
            logs = [f for f in os.listdir(mdm_latest_dir) if f.startswith('train_') and f.endswith('.log')]
            logs.sort()
            if logs:
                latest_log = os.path.join(mdm_latest_dir, logs[-1])
                with open(latest_log, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                import re
                for line in reversed(lines):
                    if 'Validation accuracy' in line:
                        m = re.findall(r"([0-9]*\.[0-9]+|[0-9]+)", line)
                        if m:
                            try:
                                mdm_val_acc = float(m[-1])
                            except Exception:
                                pass
                        break
        except Exception:
            mdm_val_acc = None

        # Adapt metadata
        mdm_adapt_id = _mf_latest_adapt_log_id(mdm_latest_dir) or ''
        mdm_adapt_cfg_path = os.path.join(mdm_latest_dir, "mdm_adapt_config.json")
        mdm_adapt_cfg_obj = _load_json(mdm_adapt_cfg_path) or {}
        mdm_target_dates_local = mdm_adapt_cfg_obj.get('target_dates', [s for s in args.target_dates.split(',') if s])
        # Defaults; will be overridden by values parsed from actual adapt log if present
        # Use runner's --alignment default/value instead of hardcoding
        adapt_alignment = mdm_adapt_cfg_obj.get('alignment', args.alignment)
        buffer_size = mdm_adapt_cfg_obj.get('buffer_size', 32)
        buffer_weighting = mdm_adapt_cfg_obj.get('buffer_weighting', 'uniform')
        buffer_decay = mdm_adapt_cfg_obj.get('buffer_decay', 2.0)
        # Prefer reading the exact superparameters used from latest adapt_*.log
        try:
            logs = [f for f in os.listdir(mdm_latest_dir) if f.startswith('adapt_') and f.endswith('.log')]
            logs.sort()
            if logs:
                latest_log = os.path.join(mdm_latest_dir, logs[-1])
                with open(latest_log, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                start_idx = None
                for i, line in enumerate(lines):
                    if 'Arguments:' in line:
                        start_idx = i
                        break
                if start_idx is not None:
                    json_lines = []
                    opened = False
                    braces = 0
                    for line in lines[start_idx:]:
                        if not opened:
                            if '{' in line:
                                opened = True
                                braces += line.count('{')
                            else:
                                continue
                        json_lines.append(line[line.find('{') if '{' in line else 0:])
                        braces += line.count('{')
                        braces -= line.count('}')
                        if opened and braces <= 0:
                            break
                    json_text = ''.join(json_lines)
                    import json as _json
                    try:
                        parsed = _json.loads(json_text)
                        adapt_alignment = parsed.get('alignment', adapt_alignment)
                        buffer_size = parsed.get('buffer_size', buffer_size)
                        buffer_weighting = parsed.get('buffer_weighting', buffer_weighting)
                        buffer_decay = parsed.get('buffer_decay', buffer_decay)
                    except Exception:
                        pass
        except Exception:
            pass

        # Per-target final accuracies
        mmdd_cols_total = {
            "0326": "0326 accuracy",
            "0327": "0327 accuracy",
            "0329": "0329 accuracy",
            "0331": "0331 accuracy",
            "0401": "0401 accuracy",
            "E": "E accuracy",
            "0710": "0710 accuracy",
            "0711": "0711 accuracy",
            "0804": "0804 accuracy",
            "0814": "0814 accuracy",
            "0825": "0825 accuracy",
        }
        per_total = {}
        total_vals = []
        for td in mdm_target_dates_local:
            mmdd = _mmdd_from_date(td)
            if not mmdd and str(td).upper() == 'E':
                mmdd = 'E'
            if not mmdd:
                continue
            acc = _get_final_accuracy(mdm_latest_dir, td)
            # Coerce to python float when available to ensure proper averaging and XLSX writing
            acc_float = float(acc) if acc is not None else None
            col_name = mmdd_cols_total.get(mmdd)
            if col_name:
                per_total[col_name] = acc_float
            if acc_float is not None:
                total_vals.append(acc_float)
        avg_total = (sum(total_vals) / len(total_vals)) if total_vals else None

        mdm_row = [
            mdm_train_id,
            mdm_data_version,
            ','.join(mdm_source_dates_local),
            mdm_source_num,
            n_class,
            mdm_seed_local,
            mdm_alignment_train,
            mdm_val_acc,
            mdm_adapt_id,
            ','.join(mdm_target_dates_local),
            adapt_alignment,
            buffer_size,
            buffer_weighting,
            buffer_decay,
            per_total.get("0326 accuracy", None),
            per_total.get("0327 accuracy", None),
            per_total.get("0329 accuracy", None),
            per_total.get("0331 accuracy", None),
            per_total.get("0401 accuracy", None),
            per_total.get("E accuracy", None),
            per_total.get("0710 accuracy", None),
            per_total.get("0711 accuracy", None),
            per_total.get("0804 accuracy", None),
            per_total.get("0814 accuracy", None),
            per_total.get("0825 accuracy", None),
            avg_total,
        ]

        mdm_report_path = os.path.join(args.mdm_output_path, "mdm_summary.xlsx")
        try:
            if os.path.isfile(mdm_report_path):
                from openpyxl import load_workbook
                wb = load_workbook(mdm_report_path)
                ws = wb.active
                # If the sheet is empty, write headers
                if ws.max_row == 1 and all((cell.value is None) for cell in ws[1]):
                    ws.append(mdm_headers)
                ws.append(mdm_row)
                wb.save(mdm_report_path)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(mdm_headers)
                ws.append(mdm_row)
                wb.save(mdm_report_path)
            print(f"[INFO] Wrote MDM summary to {mdm_report_path}")
        except Exception as e:
            print(f"[WARN] Failed to write MDM summary: {e}")
    except Exception as e:
        print(f"[WARN] MDM summary block failed: {e}")

    # 12) Write resemble XLSX report (adapt-only summary)
    try:
        # Refresh latest resemble dir in case newer runs occurred after earlier detection
        try:
            res_latest_dir = find_latest_dir(args.resemble_output_path)
            print(f"[INFO] (refresh) Latest resemble directory: {res_latest_dir}")
        except Exception:
            pass
        res_headers = [
            "Adapt ID",
            "Data Version",
            "MDM class means",
            "OTTA Checkpoint",
            "Source date",
            "Source domain number",
            "n_class",
            "seed",
            "alignment",
            "Target date",
            "buffer_size",
            "buffer_weighting",
            "buffer_decay",
            "0326 accuracy\n (deep)",
            "0326 accuracy\n (spd)",
            "0326 accuracy",
            "0327 accuracy\n (deep)",
            "0327 accuracy\n (spd)",
            "0327 accuracy",
            "0329 accuracy\n (deep)",
            "0329 accuracy\n (spd)",
            "0329 accuracy",
            "0331 accuracy\n (deep)",
            "0331 accuracy\n (spd)",
            "0331 accuracy",
            "0401 accuracy\n (deep)",
            "0401 accuracy\n (spd)",
            "0401 accuracy",
            "E accuracy",
            "0710 accuracy\n (deep)",
            "0710 accuracy\n (spd)",
            "0710 accuracy",
            "0711 accuracy\n (deep)",
            "0711 accuracy\n (spd)",
            "0711 accuracy",
            "0804 accuracy\n (deep)",
            "0804 accuracy\n (spd)",
            "0804 accuracy",
            "0814 accuracy\n (deep)",
            "0814 accuracy\n (spd)",
            "0814 accuracy",
            "0825 accuracy\n (deep)",
            "0825 accuracy\n (spd)",
            "0825 accuracy",
            "Avg test accuracy\n(Deep)",
            "Avg test accuracy\n(SPD)",
            "Avg test accuracy",
        ]

        # Adapt ID from latest adapt_*.log
        res_adapt_id = _mf_latest_adapt_log_id(res_latest_dir) or ''

        # MDM class means and OTTA checkpoint roots
        res_mdm_root = mdm_latest_dir
        res_checkpoint_root = latest_dir

        # Source dates and number
        if 'mdm_source_dates' in locals():
            res_source_dates = mdm_source_dates
        else:
            sd = train_args_obj.get("source_dates", []) or []
            if isinstance(sd, str):
                res_source_dates = [s for s in sd.split(',') if s]
            elif isinstance(sd, list):
                res_source_dates = sd
            else:
                res_source_dates = []
        res_source_num = len(res_source_dates)
        res_first_src = res_source_dates[0] if res_source_dates else ''

        # n_class from MDM class means
        res_n_class = ''
        cm_npz_path = os.path.join(res_mdm_root, f"class_means_source_{res_first_src}.npz") if res_first_src else ''
        if cm_npz_path and os.path.isfile(cm_npz_path):
            try:
                with np.load(cm_npz_path) as npz:
                    cm = npz.get('class_means', None)
                    if cm is not None:
                        res_n_class = cm.shape[0]
            except Exception:
                res_n_class = ''

        # Seed
        res_seed = args.seed if args.seed is not None else train_args_obj.get("seed", "")

        # Data version
        res_data_version = args.data_path or train_args_obj.get("data_path", "")

        # Parse resemble adapt config to get target dates and possible params
        resemble_cfg_path = os.path.join(latest_dir, "resemble_adapt_config.json")
        resemble_cfg_obj = _load_json(resemble_cfg_path) or {}
        res_target_dates = resemble_cfg_obj.get('target_dates', mdm_target_dates if 'mdm_target_dates' in locals() else [s for s in args.target_dates.split(',') if s])

        # Default alignment and buffer params; try to parse from latest adapt log if available
        # Use runner's --alignment default/value
        res_alignment = args.alignment
        res_buffer_size = 32
        res_buffer_weighting = 'uniform'
        res_buffer_decay = 2.0
        # Attempt to parse args JSON from latest adapt log
        try:
            logs = [f for f in os.listdir(res_latest_dir) if f.startswith('adapt_') and f.endswith('.log')]
            logs.sort()
            if logs:
                latest_log = os.path.join(res_latest_dir, logs[-1])
                with open(latest_log, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                # Find line starting with "Arguments:" and capture JSON block
                start_idx = None
                for i, line in enumerate(lines):
                    if 'Arguments:' in line:
                        start_idx = i
                        break
                if start_idx is not None:
                    # Collect until matching closing brace of JSON
                    json_lines = []
                    opened = False
                    braces = 0
                    for line in lines[start_idx:]:
                        # Find first '{' to start
                        if not opened:
                            if '{' in line:
                                opened = True
                                braces += line.count('{')
                            else:
                                continue
                        json_lines.append(line[line.find('{') if '{' in line else 0:])
                        braces += line.count('{')
                        braces -= line.count('}')
                        if opened and braces <= 0:
                            break
                    json_text = ''.join(json_lines)
                    import json as _json
                    try:
                        parsed = _json.loads(json_text)
                        res_alignment = parsed.get('alignment', res_alignment)
                        res_buffer_size = parsed.get('buffer_size', res_buffer_size)
                        res_buffer_weighting = parsed.get('buffer_weighting', res_buffer_weighting)
                        res_buffer_decay = parsed.get('buffer_decay', res_buffer_decay)
                    except Exception:
                        pass
        except Exception:
            pass

        # Per-target deep/SPD/total final accuracies from resemble adapt results
        mmdd_cols_deep = {
            "0326": "0326 accuracy\n (deep)",
            "0327": "0327 accuracy\n (deep)",
            "0329": "0329 accuracy\n (deep)",
            "0331": "0331 accuracy\n (deep)",
            "0401": "0401 accuracy\n (deep)",
            "0710": "0710 accuracy\n (deep)",
            "0711": "0711 accuracy\n (deep)",
            "0804": "0804 accuracy\n (deep)",
            "0814": "0814 accuracy\n (deep)",
            "0825": "0825 accuracy\n (deep)",
        }
        mmdd_cols_spd = {
            "0326": "0326 accuracy\n (spd)",
            "0327": "0327 accuracy\n (spd)",
            "0329": "0329 accuracy\n (spd)",
            "0331": "0331 accuracy\n (spd)",
            "0401": "0401 accuracy\n (spd)",
            "0710": "0710 accuracy\n (spd)",
            "0711": "0711 accuracy\n (spd)",
            "0804": "0804 accuracy\n (spd)",
            "0814": "0814 accuracy\n (spd)",
            "0825": "0825 accuracy\n (spd)",
        }
        mmdd_cols_total = {
            "0326": "0326 accuracy",
            "0327": "0327 accuracy",
            "0329": "0329 accuracy",
            "0331": "0331 accuracy",
            "0401": "0401 accuracy",
            "E": "E accuracy",
            "0710": "0710 accuracy",
            "0711": "0711 accuracy",
            "0804": "0804 accuracy",
            "0814": "0814 accuracy",
            "0825": "0825 accuracy",
        }
        per_deep = {}
        per_spd = {}
        per_total = {}
        deep_vals = []
        spd_vals = []
        total_vals = []
        for td in res_target_dates:
            # Prefer adaptation_results_{date}.npz; fallback to results_{date}.npz if needed
            npz_path = os.path.join(res_latest_dir, f"adaptation_results_{td}.npz")
            if not os.path.isfile(npz_path):
                alt = os.path.join(res_latest_dir, f"results_{td}.npz")
                if os.path.isfile(alt):
                    npz_path = alt
            deep_final = _npz_scalar(npz_path, "deep_accuracy")
            spd_final = _npz_scalar(npz_path, "spd_accuracy")
            total_final = _npz_scalar(npz_path, "accuracy")
            mmdd = _mmdd_from_date(td)
            if not mmdd and str(td).upper() == 'E':
                mmdd = 'E'
            if mmdd:
                deep_col = mmdd_cols_deep.get(mmdd)
                spd_col = mmdd_cols_spd.get(mmdd)
                tot_col = mmdd_cols_total.get(mmdd)
                if deep_col:
                    per_deep[deep_col] = deep_final
                if spd_col:
                    per_spd[spd_col] = spd_final
                if tot_col:
                    per_total[tot_col] = total_final
            if isinstance(deep_final, (int, float)):
                deep_vals.append(deep_final)
            if isinstance(spd_final, (int, float)):
                spd_vals.append(spd_final)
            if isinstance(total_final, (int, float)):
                total_vals.append(total_final)

        avg_deep = sum(deep_vals) / len(deep_vals) if deep_vals else None
        avg_spd = sum(spd_vals) / len(spd_vals) if spd_vals else None
        avg_total = sum(total_vals) / len(total_vals) if total_vals else None

        res_row = [
            res_adapt_id,
            res_data_version,
            res_mdm_root,
            res_checkpoint_root,
            ','.join(res_source_dates),
            res_source_num,
            res_n_class,
            res_seed,
            res_alignment,
            ','.join(res_target_dates),
            res_buffer_size,
            res_buffer_weighting,
            res_buffer_decay,
            per_deep.get("0326 accuracy\n (deep)", None),
            per_spd.get("0326 accuracy\n (spd)", None),
            per_total.get("0326 accuracy", None),
            per_deep.get("0327 accuracy\n (deep)", None),
            per_spd.get("0327 accuracy\n (spd)", None),
            per_total.get("0327 accuracy", None),
            per_deep.get("0329 accuracy\n (deep)", None),
            per_spd.get("0329 accuracy\n (spd)", None),
            per_total.get("0329 accuracy", None),
            per_deep.get("0331 accuracy\n (deep)", None),
            per_spd.get("0331 accuracy\n (spd)", None),
            per_total.get("0331 accuracy", None),
            per_deep.get("0401 accuracy\n (deep)", None),
            per_spd.get("0401 accuracy\n (spd)", None),
            per_total.get("0401 accuracy", None),
            per_total.get("E accuracy", None),
            per_deep.get("0710 accuracy\n (deep)", None),
            per_spd.get("0710 accuracy\n (spd)", None),
            per_total.get("0710 accuracy", None),
            per_deep.get("0711 accuracy\n (deep)", None),
            per_spd.get("0711 accuracy\n (spd)", None),
            per_total.get("0711 accuracy", None),
            per_deep.get("0804 accuracy\n (deep)", None),
            per_spd.get("0804 accuracy\n (spd)", None),
            per_total.get("0804 accuracy", None),
            per_deep.get("0814 accuracy\n (deep)", None),
            per_spd.get("0814 accuracy\n (spd)", None),
            per_total.get("0814 accuracy", None),
            per_deep.get("0825 accuracy\n (deep)", None),
            per_spd.get("0825 accuracy\n (spd)", None),
            per_total.get("0825 accuracy", None),
            avg_deep,
            avg_spd,
            avg_total,
        ]

        res_report_path = os.path.join(args.resemble_output_path, "resemble_summary.xlsx")
        try:
            from openpyxl import load_workbook, Workbook
            if os.path.isfile(res_report_path):
                wb = load_workbook(res_report_path)
                ws = wb.active
                if ws.max_row == 1 and all((cell.value is None) for cell in ws[1]):
                    ws.append(res_headers)
                ws.append(res_row)
                wb.save(res_report_path)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(res_headers)
                ws.append(res_row)
                wb.save(res_report_path)
            print(f"[INFO] Wrote resemble summary to {res_report_path}")
        except Exception as e:
            print(f"[WARN] Failed to write resemble summary: {e}")
    except Exception as e:
        print(f"[WARN] Resemble summary block failed: {e}")

    # 13) Write consolidated XLSX report for train/adapt parameters & results
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:
        Workbook = None
        load_workbook = None
    report_headers = [
        "Train ID",
        "Data Version",
        "Source date",
        "Alignment",
        "Loss",
        "epochs",
        "batch_size",
        "lr",
        "weight_decay",
        "n_temporal_filters",
        "temp_filter_length_inp",
        "spatial_expansion",
        "pool_length_inp",
        "pool_stride_inp",
        "dropout_inp",
        "ch_dim",
        "temp_filter_length",
        "pool_length",
        "pool_stride",
        "dropout",
        "use_feedforward",
        "Centered",
        "Validation accuracy",
        "Adapt ID",
        "Target date",
        "update",
        "Alignment",
        "buffer_size",
        "buffer_weighting",
        "buffer_decay",
        "adapt_lr",
        "adapt_weight_decay",
        "0326 accuracy",
        "0327 accuracy",
        "0329 accuracy",
        "0331 accuracy",
        "0401 accuracy",
        "E accuracy",
        "0710 accuracy",
        "0711 accuracy",
        "0804 accuracy",
        "0814 accuracy",
        "0825 accuracy",
        "Avg test accuracy",
    ]

    # Prepare train fields
    train_id = os.path.basename(latest_dir)
    data_version = args.data_path or train_args_obj.get("data_path", "")
    src_dates_str = ','.join(real_source_dates) if 'real_source_dates' in locals() and real_source_dates else ''
    train_alignment = train_args_obj.get("alignment", "")
    train_loss = train_args_obj.get("criterion", "")
    epochs = train_args_obj.get("epochs", "")
    batch_size = train_args_obj.get("batch_size", "")
    lr = train_args_obj.get("lr", "")
    weight_decay = train_args_obj.get("weight_decay", "")
    n_temporal_filters = train_args_obj.get("n_temporal_filters", "")
    temp_filter_length_inp = train_args_obj.get("temp_filter_length_inp", "")
    spatial_expansion = train_args_obj.get("spatial_expansion", "")
    pool_length_inp = train_args_obj.get("pool_length_inp", "")
    pool_stride_inp = train_args_obj.get("pool_stride_inp", "")
    dropout_inp = train_args_obj.get("dropout_inp", "")
    ch_dim = train_args_obj.get("ch_dim", "")
    temp_filter_length = train_args_obj.get("temp_filter_length", "")
    pool_length = train_args_obj.get("pool_length", "")
    pool_stride = train_args_obj.get("pool_stride", "")
    dropout = train_args_obj.get("dropout", "")
    use_feedforward = train_args_obj.get("use_feedforward", False)
    centered = True  # using load_centered_* loaders in this pipeline
    # Best validation accuracy from training history of first source
    best_valid_acc = None
    if first_src:
        hist_path = os.path.join(latest_dir, f"training_history_{first_src}.json")
        hist_obj = _load_json(hist_path)
        if hist_obj:
            best_valid_acc = hist_obj.get("best_valid_acc", None)

    # Adapt fields
    adapt_args_tuple = _latest_adapt_args(latest_dir)
    adapt_id = ''
    adapt_update = ''
    adapt_alignment = ''
    buffer_size = ''
    buffer_weighting = ''
    buffer_decay = ''
    adapt_lr_val = ''
    adapt_wd_val = ''
    tgt_dates_list = [s for s in args.target_dates.split(",") if s]
    if adapt_args_tuple:
        adapt_args_obj, adapt_id = adapt_args_tuple
        if adapt_args_obj:
            adapt_update = adapt_args_obj.get("update", "")
            adapt_alignment = adapt_args_obj.get("alignment", "")
            buffer_size = adapt_args_obj.get("buffer_size", "")
            buffer_weighting = adapt_args_obj.get("buffer_weighting", "")
            buffer_decay = adapt_args_obj.get("buffer_decay", "")
            adapt_lr_val = adapt_args_obj.get("adapt_lr", "")
            adapt_wd_val = adapt_args_obj.get("adapt_weight_decay", "")

    # Per-date accuracies and average
    mmdd_to_col = {
        "0326": "0326 accuracy",
        "0327": "0327 accuracy",
        "0329": "0329 accuracy",
        "0331": "0331 accuracy",
        "0401": "0401 accuracy",
        "E": "E accuracy",
        "0710": "0710 accuracy",
        "0711": "0711 accuracy",
        "0804": "0804 accuracy",
        "0814": "0814 accuracy",
        "0825": "0825 accuracy",
    }
    per_date_acc = {}
    acc_values = []
    for td in tgt_dates_list:
        final_acc = _get_final_accuracy(latest_dir, td)
        if final_acc is not None:
            acc_values.append(final_acc)
            mmdd = _mmdd_from_date(td)
            if not mmdd and str(td).upper() == 'E':
                mmdd = 'E'
            if mmdd and mmdd in mmdd_to_col:
                per_date_acc[mmdd_to_col[mmdd]] = final_acc
    avg_acc = sum(acc_values) / len(acc_values) if acc_values else None

    # Build row aligned to headers
    row = [
        train_id,
        data_version,
        src_dates_str,
        train_alignment,
        train_loss,
        epochs,
        batch_size,
        lr,
        weight_decay,
        n_temporal_filters,
        temp_filter_length_inp,
        spatial_expansion,
        pool_length_inp,
        pool_stride_inp,
        dropout_inp,
        ch_dim,
        temp_filter_length,
        pool_length,
        pool_stride,
        dropout,
        str(use_feedforward),
        str(centered),
        best_valid_acc,
        adapt_id,
        ','.join(tgt_dates_list),
        adapt_update,
        adapt_alignment,
        buffer_size,
        buffer_weighting,
        buffer_decay,
        adapt_lr_val,
        adapt_wd_val,
        per_date_acc.get("0326 accuracy", None),
        per_date_acc.get("0327 accuracy", None),
        per_date_acc.get("0329 accuracy", None),
        per_date_acc.get("0331 accuracy", None),
        per_date_acc.get("0401 accuracy", None),
        per_date_acc.get("E accuracy", None),
        per_date_acc.get("0710 accuracy", None),
        per_date_acc.get("0711 accuracy", None),
        per_date_acc.get("0804 accuracy", None),
        per_date_acc.get("0814 accuracy", None),
        per_date_acc.get("0825 accuracy", None),
        avg_acc,
    ]

    # Write or append to XLSX under the output root
    report_path = os.path.join(args.output_path, "otta_summary.xlsx")
    try:
        if Workbook is None or load_workbook is None:
            print(f"[WARN] openpyxl not available; skipped XLSX report at {report_path}")
        else:
            if os.path.isfile(report_path):
                wb = load_workbook(report_path)
                ws = wb.active
                # If header row differs, optionally rewrite (skip for simplicity)
                ws.append(row)
                wb.save(report_path)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(report_headers)
                ws.append(row)
                wb.save(report_path)
            print(f"[INFO] Wrote/updated XLSX report: {report_path}")
    except Exception as e:
        print(f"[WARN] Failed to write XLSX report: {e}")



if __name__ == "__main__":
    main()
